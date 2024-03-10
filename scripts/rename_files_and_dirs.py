import os
from typing import Union
from constants import TRANSLATE
from domain import AspectRatio, Orientation, ImageParameters, DistributedPictures
from pathlib import Path
from PIL import Image, UnidentifiedImageError

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def _format_xlsx_title(ws: Worksheet, title: list[str], extend_width: int = 0) -> None:
    """Форматирование заголовка"""
    title_font = Font(bold=True)
    title_border = Border(
        bottom=Side(border_style="thin"),
        top=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin")
    )
    title_alignment = Alignment(horizontal="center", vertical="center")
    ws.append(title)

    for cell in ws[ws.max_row]:
        cell.font = title_font
        cell.border = title_border
        cell.alignment = title_alignment

    max_length = max(len(cell_value) for cell_value in title)
    for column in ws.columns:
        ws.column_dimensions[column[0].column_letter].width = max_length + extend_width


def to_excel(images: list, columns_name: list[str], file_name: Path) -> None:
    wb = Workbook()
    ws: Worksheet = wb.active
    # TODO: Изменить название страницы
    ws.title = "Страница"
    _format_xlsx_title(ws=ws, title=columns_name)

    for item in images:
        ws.append(tuple(item.model_dump().values()))

    wb.save(file_name)


def get_extension_by_coefficient(value: Union[int, float]) -> AspectRatio:
    """
    Получить соотношение сторон в зависимости от коэффициента
    Параметры:
        value (int | float): коэффициент

    Возвращаемое значение:
        AspectRatio : соотношение сторон
    """
    if 1 <= value <= 1.1:
        return AspectRatio.FIFTY_BY_FIFTY
    elif 1.100000001 <= value <= 1.3:
        return AspectRatio.FIFTY_BY_SIXTY
    elif 1.300000001 <= value <= 1.5:
        return AspectRatio.FIFTY_BY_SEVENTY
    elif 1.500000001 <= value <= 1.7:
        return AspectRatio.FIFTY_BY_EIGHTY
    elif 1.700000001 <= value <= 1.9:
        return AspectRatio.FIFTY_BY_NINETY
    elif 1.900000001 <= value <= 5:
        return AspectRatio.FIFTY_BY_ONE_HUNDRED


def get_orientation_by_aspect_ratio(
        width_value: Union[int, float],
        height_value: Union[int, float]
) -> Orientation:
    """
    Получение ориентации в зависимости от соотношения сторон

    Параметры:
        width_value (int | float): ширина
        height_value (int | float): высота

    Возвращаемое значение:
        Orientation : ориентация
    """
    max_value = max(width_value, height_value)
    min_value = min(width_value, height_value)

    if 1 <= round(max_value / min_value, 3) < 1.1:
        return Orientation.SQUARE
    elif width_value == max_value:
        return Orientation.HORIZONTAL
    else:
        return Orientation.VERTICAL


def translate(text: str) -> str:
    """
    Перевести входящий текст

    Параметры:
        text (str): текст который необходимо перевести

    Возвращаемое значение:
        text_translate (str): переведенный текст
    """
    text_translate = ''
    for letter in text.lower():
        if letter in TRANSLATE:
            text_translate += TRANSLATE[letter]
        else:
            text_translate += letter
    return text_translate


def rename_folders_in_directory(directory: Path) -> list[Path]:
    """
    Переименовать папки в полученной директории

    Параметры:
        directory (Path): полный путь до папки, в которой необходимо переименовать содержимое

    Возвращаемое значение:
        folders_names (list[Path]): список путей до переименованных папок
    """
    folders_names = []
    for folder_name in os.listdir(directory):
        folder_name_translate = translate(folder_name).replace(' ', '-')
        path_folder_name_translate = directory / folder_name_translate
        os.rename(directory / folder_name, path_folder_name_translate)
        folders_names.append(path_folder_name_translate)
    return folders_names


def get_path(directory: Path, base_path_name: str) -> list[Path]:
    subdirectories = []

    parents = [directory]
    for name in directory.parents:
        parents.append(name)
    parents.append(directory)

    if directory.name == base_path_name:
        return [directory]

    for subdirectory in parents:
        subdirectories.append(subdirectory)
        if subdirectory.name == base_path_name:
            break

    return subdirectories[::-1]


def create_category(directory: Path, base_path_name: str) -> str:
    category = "Картины"

    for subdirectory in get_path(directory, base_path_name):
        category += f" > {subdirectory.name}"

    return category


def create_old_full_path_dir(directory: Path, base_path_name: str) -> Path:
    path = Path("/")
    for subdirectory in get_path(directory, base_path_name):
        path = path / subdirectory.name

    return path


def create_new_full_path_dir(directory: Path, base_path_name: str) -> Path:
    path = Path("/")
    for subdirectory in get_path(directory, base_path_name):
        path = path / translate(subdirectory.name)

    return path


def rename_path(images_parameters: list[ImageParameters], root_dir: Path) -> None:
    result = {}
    for path_directory in images_parameters:
        path = root_dir.parent / Path(str(path_directory.old_full_path_dir)[1:])
        path_directory_parts = path.parts
        index_root = list(path_directory_parts).index(root_dir.name)
        root = False
        old_name = Path()
        new_name = Path()
        for num in range(index_root, len(path_directory_parts)):
            for direct in path_directory_parts[:num + 1]:
                old_name = old_name.joinpath(direct)

                if root:
                    new_name = new_name.joinpath(translate(direct))
                else:
                    new_name = new_name.joinpath(direct)

                if direct == root_dir.name:
                    root = True

            if old_name not in result:
                result[old_name] = new_name

    sorted_paths = sorted(result, key=lambda x: len(str(x)))[::-1]

    for old_path in sorted_paths:
        os.rename(old_path, old_path.parent / result[old_path].name)

    os.rename(root_dir, root_dir.parent / translate(root_dir.name))


def run_fast_scandir(directory: Path, article: int, base_directory_name: str, log: list[str]):
    count = 1
    sub_folders = []
    images_parameters = []

    old_name_dir = directory.name
    new_name_dir = translate(directory.name)
    old_full_path_dir = create_old_full_path_dir(directory, base_directory_name)
    new_full_path_dir = create_new_full_path_dir(directory, base_directory_name)

    # Перебираем все папки в директории
    for file in directory.iterdir():
        category = create_category(directory, base_directory_name)

        if file.is_dir():
            sub_folders.append(file)

        if file.is_file():
            new_img_name = f"{new_name_dir}-{count}.jpg"
            try:
                img = Image.open(file)
                width, height = img.size
                img.save(directory / new_img_name)
                coefficient = round(max(width, height) / min(width, height), 3)
                aspect_ratio = get_extension_by_coefficient(coefficient)
                orientation = get_orientation_by_aspect_ratio(width, height)
                os.remove(file)
            except (UnidentifiedImageError, OSError):
                new_img_name = f"Не удается идентифицировать файл изображения: {file}"
                log.append(new_name_dir)
                width = height = coefficient = aspect_ratio = orientation = None

            img_name = f"{directory.name} Арт.{article}"
            img_link = f"https://liss-art.ru/wp-content/uploads/img-product1" + str(new_full_path_dir) + new_img_name

            image_parameters = ImageParameters(
                old_name_dir=str(old_name_dir),
                new_name_dir=str(new_name_dir),
                new_img_name=new_img_name,
                old_full_path_dir=str(old_full_path_dir),
                new_full_path_dir=str(new_full_path_dir),
                width=width,
                height=height,
                coefficient=coefficient,
                aspect_ratio=aspect_ratio,
                orientation=orientation,
                article=article,
                category=category,
                img_link=img_link,
                img_name=img_name,
            )
            article += 1
            count += 1
            images_parameters.append(image_parameters)

    for dir in list(sub_folders):
        sf, file, article, log = run_fast_scandir(dir, article, base_directory_name, log)
        sub_folders.extend(sf)
        images_parameters.extend(file)

    return sub_folders, images_parameters, article, log


