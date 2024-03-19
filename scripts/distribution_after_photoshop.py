import os
import random
import shutil
from pathlib import Path

from openpyxl import load_workbook

import settings
from domain import DistributedPictures


def process_string(input_string, new_element) -> str:
    # Разделение строки по запятой и добавление нового элемента
    elements = input_string.split(',')
    elements.append(new_element)

    # Перемешивание списка
    if settings.is_random_links:
        random.shuffle(elements)

    # Соединение списка обратно в строку с разделителем запятой
    output_string = ', '.join(elements)

    return output_string


def run_distribution_files_in_base_path(file_excel: Path, base_path: Path):
    workbook = load_workbook(filename=file_excel)
    sheet = workbook["Страница"]

    name_path = "Распределенные_картинки_после_фотошопа"
    directory_path = file_excel.parent / name_path

    if not os.path.exists(str(directory_path)):
        os.mkdir(directory_path)

    logs = []
    distributed_pictures_dict: dict[str, DistributedPictures] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        old_img_path = Path(row[14])
        directory_parent = Path(str(old_img_path).replace(str(old_img_path.parent.parent), str(base_path)))
        old_img_path = directory_parent

        img_name = old_img_path.name.replace(".jpg", "")
        new_img_path = Path(row[4])
        images_path = sorted(old_img_path.parent.glob('*.jpg'))

        for img_path in images_path:
            if img_name in img_path.name:
                Path(directory_path / str(new_img_path)[1:]).mkdir(parents=True, exist_ok=True)
                copy_to = str(directory_path / str(new_img_path)[1:] / img_path.name)
                copy_from = img_path

                new_img_link = settings.url + str(new_img_path / img_path.name).replace('\\', "/")

                if row[2] not in distributed_pictures_dict:
                    distributed_pictures = DistributedPictures(
                            old_name_dir=row[0],
                            new_name_dir=row[1],
                            new_img_name=row[2],
                            old_full_path_dir=row[3],
                            new_full_path_dir=str(new_img_path / img_path.name),
                            width=row[5],
                            height=row[6],
                            coefficient=row[7],
                            aspect_ratio=row[8],
                            orientation=row[9],
                            article=row[10],
                            category=row[11],
                            img_link=new_img_link,
                            img_name=row[13],
                            new_path=str(old_img_path),
                        )
                    distributed_pictures_dict[row[2]] = distributed_pictures

                else:
                    distributed_pictures = distributed_pictures_dict.get(row[2])
                    distributed_pictures.img_link = process_string(distributed_pictures.img_link, new_img_link)
                    distributed_pictures.new_path += f", {str(old_img_path)}"

                shutil.copy(copy_from, copy_to)

    return distributed_pictures_dict.values(), logs
