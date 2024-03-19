from pathlib import Path

from openpyxl import load_workbook

import settings
from domain import FinalTableWithVariations


def get_variable():
    workbook = load_workbook(filename=settings.variable)
    sheet = workbook["Вариации"]
    variable = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] not in variable:
            variable[row[0]] = []

        variable[row[0]].append(row)

    return variable


def read_excel(file_excel: Path):
    workbook = load_workbook(filename=file_excel)
    sheet = workbook["Страница"]
    return sheet.iter_rows(min_row=2, values_only=True)


def run_final_table_with_variations(file_excel: Path):
    data = read_excel(file_excel)
    final_tables = []
    logs = []
    variable = get_variable()

    for row in data:
        name_img = row[3]
        parent = row[2]

        pages = []
        pages.append(row)
        vars = variable.get(row[0])

        if vars:
            for var in vars:
                pages.append(var)

        if pages:
            for page in pages:
                page = list(page)
                page[3] = name_img

                final_tables_dict = {}
                for i, key in enumerate(FinalTableWithVariations.__fields__.keys()):
                    if page[i] is not None:
                        final_tables_dict[key] = page[i]

                final_table = FinalTableWithVariations(**final_tables_dict)
                final_table.parent = parent

                final_table.article = str(final_table.parent)
                if final_table.position:
                    final_table.article = str(final_table.parent) + "-" + str(final_table.position)

                final_tables.append(final_table)

    return final_tables, logs
