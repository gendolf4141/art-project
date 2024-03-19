import os
import shutil
from pathlib import Path

from openpyxl import load_workbook

from domain import DistributedPictures, Orientation


def create_aspect_ratio_name(aspect_ratio: str, orientation: str):
    if orientation == Orientation.HORIZONTAL.value:
        return f"{aspect_ratio.replace('*', '')}-G"
    elif orientation == Orientation.VERTICAL.value:
        return f"{aspect_ratio.replace('*', '')}-V"
    elif orientation == Orientation.SQUARE.value:
        return aspect_ratio.replace('*', '')


class DistributionBySize:
    sheet_name = "Страница"
    file_excel = ""


def read_excel(file_excel: Path):
    workbook = load_workbook(filename=file_excel)
    sheet = workbook["Страница"]
    return sheet.iter_rows(min_row=2, values_only=True)


def get_unique_values(file_excel: Path):
    data = read_excel(file_excel)

    name_path = "Распределенные_картинки"
    directory_path = file_excel.parent / name_path

    if not os.path.exists(str(directory_path)):
        os.mkdir(directory_path)

    values = []
    logs = []

    for row in data:
        aspect_ratio = row[8]
        orientation = row[9]
        file_name = row[2]
        file_path = row[4]
        file_full_path = file_excel.parent / Path(file_path[1:]) / file_name

        if aspect_ratio and orientation and file_name and file_path:
            aspect_ratio_name = create_aspect_ratio_name(aspect_ratio, orientation)
            aspect_ratio_directory = directory_path / aspect_ratio_name

            if not os.path.exists(aspect_ratio_directory):
                os.mkdir(aspect_ratio_directory)

            shutil.copy(file_full_path, aspect_ratio_directory / file_name)
            new_path = str(aspect_ratio_directory / file_name)
        else:
            new_path = "Не достаточно информации для формирования пути"
            logs.append(file_full_path)

        distributed_pictures = DistributedPictures(
            old_name_dir=row[0],
            new_name_dir=row[1],
            new_img_name=row[2],
            old_full_path_dir=row[3],
            new_full_path_dir=row[4],
            width=row[5],
            height=row[6],
            coefficient=row[7],
            aspect_ratio=row[8],
            orientation=row[9],
            article=row[10],
            category=row[11],
            img_link=row[12],
            img_name=row[13],
            new_path=new_path,
        )
        values.append(distributed_pictures)

    return values, logs


def run_distribution_by_size():
    pass
