from pathlib import Path
from openpyxl import load_workbook
from domain import FinalTable
from constants import SIZE


def read_excel(file_excel: Path):
    workbook = load_workbook(filename=file_excel)
    sheet = workbook["Страница"]
    return sheet.iter_rows(min_row=2, values_only=True)


def run_final_table_without_variations(file_excel: Path):
    data = read_excel(file_excel)
    final_tables = []
    logs = []
    for row in data:
        final_table = FinalTable(
            id=row[8],
            article=int(row[10]),
            name=f"Картина {row[0]} {row[10]}",
            category=row[11],
            images=row[12],
            attribute_values_1=SIZE.get(row[8])[0],
            default_attribute_1=SIZE.get(row[8])[1],
            attribute_values_4=row[9],
        )
        final_tables.append(final_table)

    return final_tables, logs
